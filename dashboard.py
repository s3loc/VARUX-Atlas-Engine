#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VARUX ELITE ENTERPRISE SECURITY PLATFORM v6.0 - ORCHESTRATOR EDITION
© 2025 VARUX Security - Tüm Hakları Saklıdır.
Kurumsal Lisanslı Ürün - Advanced Orchestration System
"""

from flask import Flask, request, jsonify
from urllib import request as urllib_request
import dash
from dash import dcc, html, Input, Output, State, dash_table, callback_context
import dash_bootstrap_components as dbc
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json
import os
from pathlib import Path
import threading
import time
from datetime import datetime, timedelta
import random
import smtplib
import subprocess
import sys
import asyncio
import importlib.util
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from varux.core.modules import MODULE_REGISTRY

# DOĞRU OLANLAR
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

import sqlite3
import hashlib
import secrets
import io
import base64
from openpyxl import Workbook
import warnings

warnings.filterwarnings("ignore")

# ====================== ORCHESTRATOR KONFİGÜRASYONU ======================
# Modülleri dinamik olarak bulabilmek için çalışma dizinini kullanıyoruz.
# Kullanıcı isterse ``VARUX_BASE_DIR`` ortam değişkeniyle özel bir dizin
# tanımlayabilir; aksi halde bu dosyanın bulunduğu dizinin kökü kullanılır.
ENV_BASE_DIR = os.getenv("VARUX_BASE_DIR")
DEFAULT_BASE_DIR = Path(__file__).resolve().parent
VARUX_BASE_DIR = Path(ENV_BASE_DIR).expanduser() if ENV_BASE_DIR else DEFAULT_BASE_DIR
VARUX_MODULE_DIR = VARUX_BASE_DIR / "varux"

# Modül yolu bulunamazsa varsayılan çalışma dizinine geri dön ve uyarı ver.
if not VARUX_MODULE_DIR.exists():
    print(f"[WARN] VARUX modül dizini bulunamadı: {VARUX_MODULE_DIR}. Varsayılana dönülüyor.")
    VARUX_BASE_DIR = DEFAULT_BASE_DIR
    VARUX_MODULE_DIR = VARUX_BASE_DIR / "varux"

sys.path.insert(0, str(VARUX_MODULE_DIR))

# ====================== GELİŞMİŞ MODÜL YÖNETİCİSİ ======================
class AdvancedModuleOrchestrator:
    def __init__(self):
        self.modules = MODULE_REGISTRY
        self.executor = ThreadPoolExecutor(max_workers=10)
        self.active_scans = {}
        self.api_url = os.getenv("VARUX_ORCH_URL", "http://127.0.0.1:5001")
        
    def load_module(self, module_path):
        """Güvenli modül yükleme"""
        try:
            spec = importlib.util.spec_from_file_location(
                module_path.stem.replace(" ", "_"),
                str(module_path)
            )
            if spec is None:
                raise ImportError(f"Module spec could not be created for {module_path}")
                
            module = importlib.util.module_from_spec(spec)
            sys.modules[module_path.stem.replace(" ", "_")] = module
            spec.loader.exec_module(module)
            return module
        except Exception as e:
            print(f"❌ Module loading error for {module_path}: {e}")
            return None

    def normalize_output(self, raw_output, scan_type, target):
        """Ham çıktıyı standardize JSON formatına dönüştür"""
        try:
            if isinstance(raw_output, dict):
                # Zaten JSON formatında
                normalized = raw_output
            elif isinstance(raw_output, str):
                # String'i JSON'a çevirmeye çalış
                try:
                    normalized = json.loads(raw_output)
                except:
                    normalized = {'raw_output': raw_output}
            else:
                normalized = {'output': str(raw_output)}
            
            # Standart alanları ekle
            standardized = {
                'scan_type': scan_type,
                'target': target,
                'timestamp': datetime.now().isoformat(),
                'status': 'completed',
                'critical_findings': normalized.get('critical_findings', 0),
                'total_devices': normalized.get('total_devices', 0),
                'vulnerabilities': normalized.get('vulnerabilities', []),
                'devices': normalized.get('devices', []),
                'risk_score': normalized.get('risk_score', 0),
                'execution_time': normalized.get('execution_time', 0),
                'raw_data': normalized
            }
            
            return standardized
            
        except Exception as e:
            print(f"❌ Output normalization error: {e}")
            return {
                'scan_type': scan_type,
                'target': target,
                'timestamp': datetime.now().isoformat(),
                'status': 'error',
                'error': str(e),
                'critical_findings': 0,
                'total_devices': 0,
                'vulnerabilities': [],
                'devices': [],
                'risk_score': 0,
                'execution_time': 0
            }

    async def execute_async_module(self, module_info, *args, **kwargs):
        """Asenkron modül çalıştırma"""
        try:
            module_path = VARUX_MODULE_DIR / module_info['file']
            module = self.load_module(module_path)
            
            if not module:
                return {"error": f"Module {module_info['file']} could not be loaded"}
                
            if 'class' in module_info:
                # Sınıf tabanlı modüller
                class_obj = getattr(module, module_info['class'])()
                if module_info['async']:
                    result = await getattr(class_obj, module_info['function'])(*args, **kwargs)
                else:
                    result = getattr(class_obj, module_info['function'])(*args, **kwargs)
            else:
                # Fonksiyon tabanlı modüller
                if module_info['async']:
                    result = await getattr(module, module_info['function'])(*args, **kwargs)
                else:
                    result = getattr(module, module_info['function'])(*args, **kwargs)
                    
            return result
            
        except Exception as e:
            print(f"❌ Async module execution error: {e}")
            return {"error": str(e)}

    def execute_sync_module(self, module_info, *args, **kwargs):
        """Senkron modül çalıştırma"""
        try:
            module_path = VARUX_MODULE_DIR / module_info['file']
            module = self.load_module(module_path)
            
            if not module:
                return {"error": f"Module {module_info['file']} could not be loaded"}
                
            if 'class' in module_info:
                # Sınıf tabanlı modüller
                class_obj = getattr(module, module_info['class'])()
                result = getattr(class_obj, module_info['function'])(*args, **kwargs)
            else:
                # Fonksiyon tabanlı modüller
                result = getattr(module, module_info['function'])(*args, **kwargs)
                
            return result
            
        except Exception as e:
            print(f"❌ Sync module execution error: {e}")
            return {"error": str(e)}

    def analyze_target(self, scan_type, target, user_data):
        """Orchestrator API üzerinden tarama başlat."""

        payload = {"module": scan_type, "payload": {"target": target}}
        try:
            response = requests.post(f"{self.api_url}/api/tasks", json=payload, timeout=10)
            response.raise_for_status()
            data = response.json()
            scan_id = data.get("job_id")
            self.active_scans[scan_id] = {
                'status': 'PENDING',
                'start_time': datetime.now(),
                'scan_type': scan_type,
                'target': target,
                'user': user_data.get('username', 'dashboard'),
            }
            return scan_id
        except Exception as exc:
            print(f"❌ Orchestrator enqueue error: {exc}")
            return None

    def save_scan_result(self, user_data, scan_type, target, result):
        """Tarama sonucunu veritabanına kaydet"""
        try:
            conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM users WHERE username = ?', (user_data['username'],))
            user_id = cursor.fetchone()[0]
            
            cursor.execute('''
                INSERT INTO scan_history (user_id, scan_type, target, results_json, critical_findings)
                VALUES (?, ?, ?, ?, ?)
            ''', (user_id, scan_type, target, json.dumps(result), result.get('critical_findings', 0)))
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"❌ Database save error: {e}")

    def update_live_data(self, result):
        """Global veriyi güncelle"""
        try:
            if 'devices' in result and result['devices']:
                LIVE_DATA["devices"] = result['devices']
                LIVE_DATA["last_update"] = datetime.now()
                
                # İstatistikleri güncelle
                LIVE_DATA["total_devices"] = len(result['devices'])
                LIVE_DATA["high_risk_devices"] = len([d for d in result['devices'] if d.get('criticality') == 'high'])
                
            print(f"📊 Live data updated: {len(result.get('devices', []))} devices")
            
        except Exception as e:
            print(f"❌ Live data update error: {e}")

    def send_critical_alert(self, user_data, result):
        """Kritik bulgu bildirimi gönder"""
        try:
            if not user_data.get('email_notifications', True):
                return
                
            critical_count = result.get('critical_findings', 0)
            if critical_count > 0:
                email_service.send_alert(
                    user_data['email'],
                    f"Critical Security Findings - {result['scan_type']}",
                    f"Scan on {result['target']} found {critical_count} critical issues. "
                    f"Total devices: {result.get('total_devices', 0)}. "
                    f"Risk score: {result.get('risk_score', 0)}",
                    attachment_data=None
                )
                
        except Exception as e:
            print(f"❌ Alert sending error: {e}")

    def get_scan_status(self, scan_id):
        """Tarama durumunu getir"""

        if not scan_id:
            return {'status': 'unknown'}

        try:
            response = requests.get(f"{self.api_url}/api/tasks/{scan_id}", timeout=5)
            response.raise_for_status()
            data = response.json()
            status = data.get('status', 'unknown').lower()
            self.active_scans[scan_id] = data
            return {'status': status, 'result': data.get('result'), 'meta': data.get('meta')}
        except Exception as exc:
            return {'status': 'error', 'error': str(exc)}

    def invoke_assistant(self, prompt, context=None):
        """OpenAI tabanlı kod asistanını çalıştır."""

        try:
            module_info = self.modules.get('ai_assistant')
            if not module_info:
                return {"error": "Asistan modülü kayıtlı değil."}

            if module_info.get('async'):
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    return loop.run_until_complete(
                        self.execute_async_module(module_info, prompt, context)
                    )
                finally:
                    loop.close()

            return self.execute_sync_module(module_info, prompt, context)

        except Exception as e:
            print(f"❌ Assistant invocation error: {e}")
            return {"error": str(e)}

# ====================== VERİTABANI KURULUMU ======================
def init_database():
    try:
        conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
        cursor = conn.cursor()
        
        # Kullanıcılar tablosu
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                company TEXT,
                role TEXT DEFAULT 'user',
                api_key TEXT UNIQUE,
                email_notifications BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                is_active BOOLEAN DEFAULT 1
            )
        ''')
        
        # Tarama geçmişi tablosu
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS scan_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                scan_type TEXT NOT NULL,
                target TEXT NOT NULL,
                status TEXT DEFAULT 'completed',
                results_json TEXT,
                critical_findings INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # API erişim logları
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS api_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                api_key TEXT,
                endpoint TEXT,
                ip_address TEXT,
                user_agent TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Varsayılan admin kullanıcı
        admin_hash = hashlib.sha256("admin123".encode()).hexdigest()
        cursor.execute('''
            INSERT OR IGNORE INTO users (username, email, password_hash, role, company, api_key)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', ('admin', 'admin@varux.com', admin_hash, 'admin', 'VARUX Security', secrets.token_hex(32)))
        
        conn.commit()
        conn.close()
        print("✅ Veritabanı başarıyla başlatıldı")
    except Exception as e:
        print(f"❌ Veritabanı başlatma hatası: {e}")

# Veritabanını başlat
init_database()

# ====================== E-POSTA SERVİSİ ======================
class EmailService:
    def __init__(self):
        self.smtp_server = "smtp.gmail.com"
        self.smtp_port = 587
        # Bu bilgileri environment variables'dan almanız önerilir
        self.sender_email = "security@varux.com"
        self.sender_password = "your-email-password"
    
    def send_alert(self, recipient_email, subject, message, attachment_data=None):
        """Kritik bulgu e-postası gönder"""
        try:
            msg = MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = recipient_email
            msg['Subject'] = f"🔒 VARUX Güvenlik Uyarısı: {subject}"
            
            # HTML içerik
            html_content = f"""
            <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 20px; }}
                    .header {{ background: linear-gradient(135deg, #1a1a2e, #16213e); padding: 20px; color: white; }}
                    .content {{ padding: 20px; background: #f5f5f5; }}
                    .alert {{ background: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; margin: 10px 0; }}
                    .footer {{ background: #343a40; color: white; padding: 10px; text-align: center; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <h2>VARUX Elite Security Platform</h2>
                    <p>Güvenlik Uyarı Sistemi</p>
                </div>
                <div class="content">
                    <div class="alert">
                        <h3>⚠️ {subject}</h3>
                        <p>{message}</p>
                    </div>
                    <p><strong>Zaman:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p>Bu e-posta otomatik olarak gönderilmiştir. Lütfen yanıtlamayınız.</p>
                </div>
                <div class="footer">
                    <p>© 2025 VARUX Security - Tüm Hakları Saklıdır.</p>
                </div>
            </body>
            </html>
            """
            
            msg.attach(MIMEText(html_content, 'html'))
            
            # Eklenti varsa ekle
            if attachment_data:
                attachment = MIMEApplication(attachment_data)
                attachment.add_header('Content-Disposition', 'attachment', filename='security_report.pdf')
                msg.attach(attachment)
            
            # E-posta gönder
            server = smtplib.SMTP(self.smtp_server, self.smtp_port)
            server.starttls()
            server.login(self.sender_email, self.sender_password)
            server.send_message(msg)
            server.quit()
            
            print(f"✅ E-posta gönderildi: {recipient_email}")
            return True
            
        except Exception as e:
            print(f"❌ E-posta gönderilemedi: {e}")
            return False

# ====================== RAPOR SERVİSİ ======================
class ReportService:
    @staticmethod
    def generate_excel_report(data, scan_info):
        """Excel raporu oluştur"""
        wb = Workbook()
        
        # Cihazlar sayfası
        ws_devices = wb.active
        ws_devices.title = "Keşfedilen Cihazlar"
        
        if data["devices"]:
            df_devices = pd.DataFrame(data["devices"])
            for col_num, column_title in enumerate(df_devices.columns, 1):
                ws_devices.cell(row=1, column=col_num, value=column_title)
            
            for row_num, row_data in enumerate(df_devices.values, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws_devices.cell(row=row_num, column=col_num, value=cell_value)
        
        # Özet sayfası
        ws_summary = wb.create_sheet("Tarama Özeti")
        summary_data = [
            ["Tarama Türü", scan_info.get('scan_type', 'Bilinmiyor')],
            ["Hedef", scan_info.get('target', 'Bilinmiyor')],
            ["Toplam Cihaz", len(data["devices"])],
            ["Yüksek Riskli", len([d for d in data["devices"] if d.get('criticality') == 'high'])],
            ["Tarama Zamanı", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Oluşturan", scan_info.get('username', 'Sistem')]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            ws_summary.cell(row=row_num, column=1, value=row_data[0])
            ws_summary.cell(row=row_num, column=2, value=row_data[1])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def generate_pdf_report(data, scan_info):
        """Basit PDF raporu (gerçek uygulamada WeasyPrint veya başka kütüphane kullanın)"""
        # Bu örnekte basit HTML çıktısı döndürüyoruz
        html_content = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial; margin: 20px; }}
                .header {{ background: #1a1a2e; color: white; padding: 20px; }}
                .summary {{ background: #f5f5f5; padding: 15px; margin: 10px 0; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #343a40; color: white; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>VARUX Security Raporu</h1>
                <p>Kurumsal Güvenlik Tarama Sonuçları</p>
            </div>
            
            <div class="summary">
                <h3>Tarama Özeti</h3>
                <p><strong>Tarama Türü:</strong> {scan_info.get('scan_type', 'Bilinmiyor')}</p>
                <p><strong>Hedef:</strong> {scan_info.get('target', 'Bilinmiyor')}</p>
                <p><strong>Toplam Cihaz:</strong> {len(data["devices"])}</p>
                <p><strong>Oluşturulma Zamanı:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            
            <h3>Keşfedilen Cihazlar</h3>
            <table>
                <tr>
                    <th>IP Adresi</th>
                    <th>MAC Adresi</th>
                    <th>Üretici</th>
                    <th>Protokoller</th>
                    <th>Kritiklik</th>
                </tr>
        """
        
        for device in data["devices"]:
            protocols = device.get('protocols', 'N/A')
            if isinstance(protocols, list):
                protocols = ', '.join(protocols)
            
            html_content += f"""
                <tr>
                    <td>{device.get('ip', 'N/A')}</td>
                    <td>{device.get('mac', 'N/A')}</td>
                    <td>{device.get('vendor', 'N/A')}</td>
                    <td>{protocols}</td>
                    <td>{device.get('criticality', 'N/A')}</td>
                </tr>
            """
        
        html_content += """
            </table>
            <div style="margin-top: 30px; text-align: center; color: #666;">
                <p>© 2025 VARUX Security - Bu rapor gizlidir ve yalnızca yetkili personel ile paylaşılabilir.</p>
            </div>
        </body>
        </html>
        """
        
        # Gerçek uygulamada HTML'den PDF'e dönüşüm yapılacak
        return html_content.encode()

# ====================== API SERVİSİ ======================
class APIService:
    @staticmethod
    def validate_api_key(api_key):
        """API anahtarını doğrula"""
        try:
            conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute('SELECT username, role FROM users WHERE api_key = ? AND is_active = 1', (api_key,))
            user = cursor.fetchone()
            conn.close()
            return user
        except Exception as e:
            print(f"API anahtarı doğrulama hatası: {e}")
            return None
    
    @staticmethod
    def log_api_request(api_key, endpoint, ip_address, user_agent):
        """API isteğini logla"""
        try:
            conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO api_logs (api_key, endpoint, ip_address, user_agent)
                VALUES (?, ?, ?, ?)
            ''', (api_key, endpoint, ip_address, user_agent))
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"API loglama hatası: {e}")

# ====================== DASH UYGULAMASI ======================
server = Flask(__name__)
server.secret_key = secrets.token_hex(32)
app = dash.Dash(
    __name__, 
    title="VARUX Enterprise Security Platform v6.0 - Orchestrator",
    external_stylesheets=[dbc.themes.DARKLY],
    suppress_callback_exceptions=True,
    server=server
)

# Global veri - ORCHESTRATOR ENTEGRASYONLU
LIVE_DATA = {
    "devices": [
        {
            "ip": "192.168.1.10",
            "mac": "00:1B:44:11:3A:B7",
            "vendor": "Siemens",
            "protocols": "Modbus, S7",
            "criticality": "high",
            "last_seen": "14:30:22"
        },
        {
            "ip": "192.168.1.15", 
            "mac": "00:1C:42:21:4C:D8",
            "vendor": "Rockwell",
            "protocols": "EtherNet/IP",
            "criticality": "medium",
            "last_seen": "14:29:15"
        },
        {
            "ip": "192.168.1.20",
            "mac": "00:1D:43:31:5E:E9",
            "vendor": "Schneider",
            "protocols": "Modbus, OPC-UA",
            "criticality": "low",
            "last_seen": "14:28:45"
        }
    ],
    "topology": {"nodes": [], "links": []},
    "vulnerabilities": [],
    "timeline": [],
    "protocol_dist": [],
    "vendor_dist": [],
    "risk_scores": [],
    "last_update": datetime.now(),
    "current_user": None,
    "total_devices": 3,
    "high_risk_devices": 1
}

# Servisler
email_service = EmailService()
report_service = ReportService()
api_service = APIService()
orchestrator = AdvancedModuleOrchestrator()

# ====================== MODERN KURUMSAL ARAYÜZ ======================
app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    dcc.Store(id='user-store', data=None),
    dcc.Store(id='scan-data-store'),
    dcc.Store(id='active-scan-id', data=None),
    
    # Ana içerik
    html.Div(id='page-content'),
    
    # Gizli indirme bileşenleri
    dcc.Download(id="download-excel"),
    dcc.Download(id="download-pdf"),
    dcc.Interval(id='interval-component', interval=5*1000, n_intervals=0),
    dcc.Interval(id='scan-status-interval', interval=2*1000, n_intervals=0),
])

# ====================== GİRİŞ SAYFASI ======================
login_layout = dbc.Container([
    dbc.Row([
        dbc.Col([
            html.Div([
                html.H1("VARUX Enterprise Platform v6.0", className="text-center mb-4", 
                       style={'color': '#00FF99', 'fontWeight': '800', 'fontSize': '2.5rem', 'textShadow': '0 0 10px rgba(0,255,153,0.5)'}),
                html.P("Advanced Orchestration Security Management System", 
                      className="text-center mb-5", style={'color': '#80d0ff', 'fontSize': '1.2rem'}),
                
                dbc.Card([
                    dbc.CardBody([
                        html.H4("Sisteme Giriş", className="card-title text-center mb-4", style={'color': '#00FF99'}),
                        
                        dbc.Form([
                            dbc.Row([
                                dbc.Col([
                                    dbc.Label("Kullanıcı Adı", style={'color': '#e0e0e0'}),
                                    dbc.Input(id='login-username', type='text', placeholder='Kullanıcı adınızı girin',
                                             style={'backgroundColor': '#0F1114', 'border': '1px solid #00FF99', 'color': 'white', 'borderRadius': '8px'})
                                ], className="mb-3"),
                                
                                dbc.Col([
                                    dbc.Label("Şifre", style={'color': '#e0e0e0'}),
                                    dbc.Input(id='login-password', type='password', placeholder='Şifrenizi girin',
                                             style={'backgroundColor': '#0F1114', 'border': '1px solid #00FF99', 'color': 'white', 'borderRadius': '8px'})
                                ], className="mb-4"),
                            ]),
                            
                            dbc.Button("Giriş Yap", id='login-button', color="primary", size="lg", 
                                      className="w-100 mb-3", style={
                                          'backgroundColor': '#00FF99', 
                                          'border': 'none', 
                                          'color': '#0F1114',
                                          'fontWeight': 'bold',
                                          'borderRadius': '8px',
                                          'boxShadow': '0 0 15px rgba(0,255,153,0.3)'
                                      }),
                            
                            html.Div(id='login-alert'),
                            
                            html.Hr(style={'borderColor': '#333'}),
                            
                            html.P("Demo Giriş Bilgileri:", className="text-center small", style={'color': '#aaa'}),
                            html.P("Kullanıcı: admin | Şifre: admin123", 
                                  className="text-center small", style={'color': '#00FF99', 'fontFamily': 'monospace', 'fontWeight': 'bold'})
                        ])
                    ], style={'backgroundColor': '#1a1a2e', 'border': '1px solid #00FF99', 'borderRadius': '12px'})
                ], style={'maxWidth': '400px', 'margin': '0 auto', 'boxShadow': '0 0 30px rgba(0,255,153,0.2)'})
            ], style={'padding': '60px 20px'})
        ])
    ])
], fluid=True, style={'backgroundColor': '#0F1114', 'minHeight': '100vh', 'paddingTop': '80px'})

# ====================== ANA DASHBOARD ======================
def create_dashboard_layout():
    return html.Div([
        # Üst Navigasyon
        dbc.Navbar([
            dbc.Container([
                html.A([
                    html.Img(src="/assets/logo.png", height="30", className="mr-2"),
                    html.Span("VARUX ENTERPRISE v6.0", style={
                        'color': '#00FF99', 
                        'fontWeight': 'bold', 
                        'marginLeft': '10px', 
                        'fontSize': '1.1rem',
                        'textShadow': '0 0 8px rgba(0,255,153,0.5)'
                    })
                ], href="#", style={'textDecoration': 'none'}),
                
                dbc.NavbarToggler(id="navbar-toggler"),
                
                dbc.Collapse([
                    dbc.Nav([
                        dbc.NavItem(dbc.NavLink("🏠 Dashboard", href="/", style={'fontSize': '0.9rem', 'color': '#00FF99'})),
                        dbc.NavItem(dbc.NavLink("🔧 API Yönetimi", href="/api", style={'fontSize': '0.9rem', 'color': '#00FF99'})),
                        dbc.NavItem(dbc.NavLink("⚙️ Hesap Ayarları", href="/settings", style={'fontSize': '0.9rem', 'color': '#00FF99'})),
                    ], className="ml-auto", navbar=True),
                    
                    dbc.DropdownMenu([
                        dbc.DropdownMenuItem(f"👤 {LIVE_DATA['current_user']['username'] if LIVE_DATA['current_user'] else 'Kullanıcı'}", header=True),
                        dbc.DropdownMenuItem("Hesap Ayarları", href="/settings"),
                        dbc.DropdownMenuItem(divider=True),
                        dbc.DropdownMenuItem("Çıkış Yap", id="logout-button"),
                    ], nav=True, in_navbar=True, label="Hesap", align_end=True, style={'fontSize': '0.9rem'}),
                ], id="navbar-collapse", navbar=True),
            ], fluid=True)
        ], color="dark", dark=True, sticky="top", style={
            'padding': '0.5rem 1rem', 
            'backgroundColor': '#0F1114', 
            'borderBottom': '1px solid #00FF99',
            'boxShadow': '0 2px 20px rgba(0,255,153,0.1)'
        }),
        
        # Ana İçerik
        dbc.Container([
            # İstatistik Kartları - MODERN TASARIM
            dbc.Row([
                dbc.Col(dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H5("0", id="stat-devices", className="card-title", style={
                                'color': '#00FF99', 
                                'margin': '0', 
                                'fontSize': '1.8rem',
                                'textShadow': '0 0 10px rgba(0,255,153,0.5)'
                            }),
                            html.P("Toplam Cihaz", className="card-text", style={
                                'margin': '0', 
                                'fontSize': '0.8rem', 
                                'color': '#aaa'
                            })
                        ], style={'textAlign': 'center'})
                    ], style={'padding': '1.2rem'})
                ], style={
                    'backgroundColor': '#1a1a2e', 
                    'border': '1px solid #00FF99',
                    'borderRadius': '12px',
                    'boxShadow': '0 0 15px rgba(0,255,153,0.1)'
                }), width=3, className="mb-4"),
                
                dbc.Col(dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H5("0", id="stat-high-risk", className="card-title", style={
                                'color': '#ff4444', 
                                'margin': '0', 
                                'fontSize': '1.8rem',
                                'textShadow': '0 0 10px rgba(255,68,68,0.5)'
                            }),
                            html.P("Yüksek Risk", className="card-text", style={
                                'margin': '0', 
                                'fontSize': '0.8rem', 
                                'color': '#aaa'
                            })
                        ], style={'textAlign': 'center'})
                    ], style={'padding': '1.2rem'})
                ], style={
                    'backgroundColor': '#1a1a2e', 
                    'border': '1px solid #ff4444',
                    'borderRadius': '12px',
                    'boxShadow': '0 0 15px rgba(255,68,68,0.1)'
                }), width=3, className="mb-4"),
                
                dbc.Col(dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H5("0", id="stat-protocols", className="card-title", style={
                                'color': '#ffaa00', 
                                'margin': '0', 
                                'fontSize': '1.8rem',
                                'textShadow': '0 0 10px rgba(255,170,0,0.5)'
                            }),
                            html.P("Aktif Protokol", className="card-text", style={
                                'margin': '0', 
                                'fontSize': '0.8rem', 
                                'color': '#aaa'
                            })
                        ], style={'textAlign': 'center'})
                    ], style={'padding': '1.2rem'})
                ], style={
                    'backgroundColor': '#1a1a2e', 
                    'border': '1px solid #ffaa00',
                    'borderRadius': '12px',
                    'boxShadow': '0 0 15px rgba(255,170,0,0.1)'
                }), width=3, className="mb-4"),
                
                dbc.Col(dbc.Card([
                    dbc.CardBody([
                        html.Div([
                            html.H5("-", id="stat-last-scan", className="card-title", style={
                                'color': '#00ddff', 
                                'margin': '0', 
                                'fontSize': '1.8rem',
                                'textShadow': '0 0 10px rgba(0,221,255,0.5)'
                            }),
                            html.P("Son Tarama", className="card-text", style={
                                'margin': '0', 
                                'fontSize': '0.8rem', 
                                'color': '#aaa'
                            })
                        ], style={'textAlign': 'center'})
                    ], style={'padding': '1.2rem'})
                ], style={
                    'backgroundColor': '#1a1a2e', 
                    'border': '1px solid #00ddff',
                    'borderRadius': '12px',
                    'boxShadow': '0 0 15px rgba(0,221,255,0.1)'
                }), width=3, className="mb-4"),
            ], className="mb-4"),
            
            # Tarama Kontrolleri ve Grafikler
            dbc.Row([
                # Sol Kolon - Kontroller
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H5("🎯 Gelişmiş Tarama Kontrolleri", className="mb-0", style={
                            'fontSize': '1rem', 
                            'color': '#00FF99',
                            'textShadow': '0 0 5px rgba(0,255,153,0.3)'
                        })),
                        dbc.CardBody([
                            dbc.Form([
                                dbc.Row([
                                    dbc.Col([
                                        dbc.Label("Tarama Türü", style={'fontSize': '0.85rem', 'color': '#00FF99'}),
                                        dcc.Dropdown(
                                            id='scan-type',
                                            options=[
                                                {'label': '🔍 OT/ICS Derin Keşif', 'value': 'ot_framework'},
                                                {'label': '🌐 Endüstriyel Ağ Keşfi', 'value': 'industrial_recon'},
                                                {'label': '🛡️ Web Uygulama Güvenlik', 'value': 'noxim'},
                                                {'label': '⚡ Tam Otomatik Test', 'value': 'varuxctl'},
                                                {'label': '💉 SQL Injection Tarama', 'value': 'sqlmap_wrapper'}
                                            ],
                                            placeholder="Tarama türünü seçin...",
                                            style={
                                                'backgroundColor': '#0F1114', 
                                                'color': 'white', 
                                                'fontSize': '0.85rem',
                                                'border': '1px solid #00FF99',
                                                'borderRadius': '8px'
                                            }
                                        ),
                                    ], width=6),
                                    
                                    dbc.Col([
                                        dbc.Label("Hedef", style={'fontSize': '0.85rem', 'color': '#00FF99'}),
                                        dbc.Input(
                                            id='target-input',
                                            type='text',
                                            placeholder='192.168.1.0/24, 10.0.0.1-100, http://site.com',
                                            style={
                                                'fontSize': '0.85rem',
                                                'backgroundColor': '#0F1114',
                                                'border': '1px solid #00FF99',
                                                'color': 'white',
                                                'borderRadius': '8px'
                                            }
                                        ),
                                    ], width=6),
                                ], className="mb-3"),
                                
                                dbc.Row([
                                    dbc.Col([
                                        dbc.Label("E-posta Bildirimi", style={'fontSize': '0.85rem', 'color': '#00FF99'}),
                                        dbc.Input(
                                            id='email-input',
                                            type='email',
                                            placeholder='bildirim@firma.com',
                                            style={
                                                'fontSize': '0.85rem',
                                                'backgroundColor': '#0F1114',
                                                'border': '1px solid #00FF99',
                                                'color': 'white',
                                                'borderRadius': '8px'
                                            }
                                        ),
                                        dbc.FormText("Kritik bulgularda e-posta ile bilgilendirileceksiniz.", style={'fontSize': '0.75rem', 'color': '#888'}),
                                    ], width=12),
                                ], className="mb-3"),
                                
                                dbc.Button(
                                    "🚀 Gelişmiş Taramayı Başlat", 
                                    id="start-scan", 
                                    color="success", 
                                    size="md",
                                    className="w-100",
                                    style={
                                        'backgroundColor': '#00FF99', 
                                        'border': 'none', 
                                        'fontSize': '0.9rem',
                                        'color': '#0F1114',
                                        'fontWeight': 'bold',
                                        'borderRadius': '8px',
                                        'boxShadow': '0 0 15px rgba(0,255,153,0.3)'
                                    }
                                ),
                                
                                html.Div(id='scan-status', className="mt-3", style={'fontSize': '0.85rem'}),
                                html.Div(id='scan-progress', className="mt-2"),
                            ])
                        ], style={'padding': '1.2rem', 'backgroundColor': '#1a1a2e'})
                    ], className="mb-4", style={
                        'border': '1px solid #00FF99',
                        'borderRadius': '12px',
                        'boxShadow': '0 0 20px rgba(0,255,153,0.15)'
                    }),

                    dbc.Card([
                        dbc.CardHeader(html.H5("🤖 OpenAI Güvenlik Asistanı", className="mb-0", style={
                            'fontSize': '1rem',
                            'color': '#00ddff',
                            'textShadow': '0 0 5px rgba(0,221,255,0.35)'
                        })),
                        dbc.CardBody([
                            dbc.Label("İstek veya soru", style={'fontSize': '0.85rem', 'color': '#00ddff'}),
                            dcc.Textarea(
                                id='assistant-prompt',
                                placeholder='Örn: Bulunan Modbus cihazları için hızlı sertifika sertleştirme adımları, kod öner...',
                                style={
                                    'width': '100%',
                                    'backgroundColor': '#0F1114',
                                    'border': '1px solid #00ddff',
                                    'color': 'white',
                                    'borderRadius': '8px',
                                    'minHeight': '120px',
                                    'padding': '10px',
                                    'fontSize': '0.9rem'
                                }
                            ),
                            dbc.Button(
                                "🤖 Kod öner", id='ask-assistant', color='info', className='mt-3 w-100',
                                style={
                                    'backgroundColor': '#00ddff',
                                    'border': 'none',
                                    'color': '#0F1114',
                                    'fontWeight': 'bold',
                                    'borderRadius': '8px'
                                }
                            ),
                            html.Div(id='assistant-response', className='mt-3', style={'fontSize': '0.9rem'})
                        ], style={'padding': '1.2rem', 'backgroundColor': '#1a1a2e'})
                    ], className="mb-4", style={
                        'border': '1px solid #00ddff',
                        'borderRadius': '12px',
                        'boxShadow': '0 0 20px rgba(0,221,255,0.15)'
                    }),

                    # Cihaz Tablosu
                    dbc.Card([
                        dbc.CardHeader(html.H5("📋 Keşfedilen Cihazlar", className="mb-0", style={
                            'fontSize': '1rem',
                            'color': '#00FF99',
                            'textShadow': '0 0 5px rgba(0,255,153,0.3)'
                        })),
                        dbc.CardBody([
                            dash_table.DataTable(
                                id='device-table',
                                columns=[
                                    {"name": "IP Adresi", "id": "ip"},
                                    {"name": "MAC", "id": "mac"},
                                    {"name": "Üretici", "id": "vendor"},
                                    {"name": "Protokoller", "id": "protocols"},
                                    {"name": "Risk", "id": "criticality"}
                                ],
                                page_size=8,
                                style_table={
                                    'overflowX': 'auto', 
                                    'fontSize': '0.8rem',
                                    'borderRadius': '8px'
                                },
                                style_cell={
                                    'backgroundColor': '#1a1a2e',
                                    'color': 'white',
                                    'border': '1px solid #2d2d4d',
                                    'padding': '6px',
                                    'textAlign': 'left',
                                    'overflow': 'hidden',
                                    'textOverflow': 'ellipsis',
                                    'maxWidth': 0,
                                    'minWidth': '80px'
                                },
                                style_header={
                                    'backgroundColor': '#252547',
                                    'fontWeight': 'bold',
                                    'border': '1px solid #2d2d4d',
                                    'padding': '8px'
                                },
                                style_data_conditional=[
                                    {
                                        'if': {'row_index': 'odd'},
                                        'backgroundColor': '#161629',
                                    },
                                    {
                                        'if': {'filter_query': '{criticality} = "high"'},
                                        'backgroundColor': '#441111',
                                        'color': 'white'
                                    },
                                    {
                                        'if': {'filter_query': '{criticality} = "medium"'},
                                        'backgroundColor': '#443311',
                                        'color': 'white'
                                    }
                                ],
                                css=[{
                                    'selector': '.dash-spreadsheet td div',
                                    'rule': '''
                                        line-height: 1.2;
                                        max-height: 20px; min-height: 20px; height: 20px;
                                        display: block;
                                        overflow: hidden;
                                        text-overflow: ellipsis;
                                    '''
                                }]
                            )
                        ], style={'padding': '1rem', 'backgroundColor': '#1a1a2e'})
                    ], style={
                        'border': '1px solid #00FF99',
                        'borderRadius': '12px',
                        'boxShadow': '0 0 20px rgba(0,255,153,0.15)'
                    }),
                ], width=4),
                
                # Sağ Kolon - Grafikler
                dbc.Col([
                    dbc.Row([
                        dbc.Col([
                            dbc.Card([
                                dbc.CardBody([
                                    dcc.Graph(id='network-graph', style={'height': '280px'})
                                ], style={'padding': '0.8rem', 'backgroundColor': '#1a1a2e'})
                            ], style={
                                'border': '1px solid #00FF99',
                                'borderRadius': '12px',
                                'boxShadow': '0 0 15px rgba(0,255,153,0.1)'
                            }),
                        ], width=6),
                        dbc.Col([
                            dbc.Card([
                                dbc.CardBody([
                                    dcc.Graph(id='device-pie', style={'height': '280px'})
                                ], style={'padding': '0.8rem', 'backgroundColor': '#1a1a2e'})
                            ], style={
                                'border': '1px solid #00FF99', 
                                'borderRadius': '12px',
                                'boxShadow': '0 0 15px rgba(0,255,153,0.1)'
                            }),
                        ], width=6),
                    ], className="mb-4"),
                    
                    dbc.Row([
                        dbc.Col([
                            dbc.Card([
                                dbc.CardBody([
                                    dcc.Graph(id='timeline-graph', style={'height': '280px'})
                                ], style={'padding': '0.8rem', 'backgroundColor': '#1a1a2e'})
                            ], style={
                                'border': '1px solid #00FF99',
                                'borderRadius': '12px', 
                                'boxShadow': '0 0 15px rgba(0,255,153,0.1)'
                            }),
                        ], width=6),
                        dbc.Col([
                            dbc.Card([
                                dbc.CardBody([
                                    dcc.Graph(id='criticality-bar', style={'height': '280px'})
                                ], style={'padding': '0.8rem', 'backgroundColor': '#1a1a2e'})
                            ], style={
                                'border': '1px solid #00FF99',
                                'borderRadius': '12px',
                                'boxShadow': '0 0 15px rgba(0,255,153,0.1)'
                            }),
                        ], width=6),
                    ], className="mb-4"),
                    
                    # Rapor İndirme
                    dbc.Card([
                        dbc.CardBody([
                            html.H5("📄 Gelişmiş Raporlar", className="mb-3", style={
                                'fontSize': '1rem',
                                'color': '#00FF99',
                                'textShadow': '0 0 5px rgba(0,255,153,0.3)'
                            }),
                            dbc.Row([
                                dbc.Col([
                                    dbc.Button(
                                        "📊 Excel Raporu", 
                                        id="btn-excel", 
                                        color="success", 
                                        className="w-100",
                                        size="sm",
                                        style={
                                            'fontSize': '0.8rem',
                                            'backgroundColor': '#00FF99',
                                            'border': 'none',
                                            'color': '#0F1114',
                                            'fontWeight': 'bold',
                                            'borderRadius': '6px'
                                        }
                                    ),
                                ], width=6),
                                dbc.Col([
                                    dbc.Button(
                                        "📋 PDF Raporu", 
                                        id="btn-pdf", 
                                        color="danger", 
                                        className="w-100",
                                        size="sm",
                                        style={
                                            'fontSize': '0.8rem',
                                            'backgroundColor': '#ff4444',
                                            'border': 'none',
                                            'color': 'white',
                                            'fontWeight': 'bold',
                                            'borderRadius': '6px'
                                        }
                                    ),
                                ], width=6),
                            ]),
                        ], style={'padding': '1.2rem', 'backgroundColor': '#1a1a2e'})
                    ], style={
                        'border': '1px solid #00FF99',
                        'borderRadius': '12px',
                        'boxShadow': '0 0 20px rgba(0,255,153,0.15)'
                    }),
                ], width=8),
            ]),
        ], fluid=True, style={'padding': '1.5rem', 'backgroundColor': '#0F1114'}),
    ], style={'backgroundColor': '#0F1114', 'minHeight': '100vh'})

# ====================== API YÖNETİM SAYFASI ======================
def create_api_layout():
    if not LIVE_DATA['current_user']:
        return login_layout
        
    conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
    cursor = conn.cursor()
    cursor.execute('SELECT api_key FROM users WHERE username = ?', (LIVE_DATA['current_user']['username'],))
    user_api_key = cursor.fetchone()
    conn.close()
    
    api_key_value = user_api_key[0] if user_api_key else "API anahtarı bulunamadı"
    
    return dbc.Container([
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader(html.H4("🔌 Gelişmiş API Yönetim Merkezi", className="mb-0", style={
                        'fontSize': '1.2rem',
                        'color': '#00FF99',
                        'textShadow': '0 0 5px rgba(0,255,153,0.3)'
                    })),
                    dbc.CardBody([
                        html.P("API anahtarınızı kullanarak VARUX güvenlik verilerinize programatik olarak erişebilirsiniz.", 
                              className="text-muted mb-4", style={'fontSize': '0.9rem', 'color': '#aaa'}),
                        
                        dbc.Label("API Anahtarınız:", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                        dbc.Input(
                            id='api-key-input',
                            value=api_key_value,
                            readonly=True,
                            style={
                                'fontFamily': 'monospace', 
                                'backgroundColor': '#0F1114', 
                                'fontSize': '0.85rem',
                                'border': '1px solid #00FF99',
                                'color': 'white',
                                'borderRadius': '8px'
                            }
                        ),
                        dbc.FormText("Bu anahtarınızı güvenli tutun ve paylaşmayın.", style={'fontSize': '0.8rem', 'color': '#888'}),
                        
                        dbc.Button("Yeni Anahtar Oluştur", id="generate-api-key", color="warning", className="mt-3", size="sm", style={
                            'backgroundColor': '#ffaa00',
                            'border': 'none',
                            'borderRadius': '6px'
                        }),
                        
                        html.Hr(className="my-4", style={'borderColor': '#333'}),
                        
                        html.H5("Gelişmiş API Dokümantasyonu", style={'fontSize': '1.1rem', 'color': '#00FF99'}),
                        html.P("Aşağıdaki endpoint'leri kullanarak verilerinize erişebilirsiniz:", style={'fontSize': '0.9rem', 'color': '#aaa'}),
                        
                        dbc.ListGroup([
                            dbc.ListGroupItem([
                                html.H6("GET /api/v1/devices", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                                html.P("Keşfedilen cihazları listeler", className="text-muted small mb-0", style={'fontSize': '0.8rem'}),
                                html.Code("Authorization: Bearer YOUR_API_KEY", className="small", style={'fontSize': '0.75rem', 'color': '#00FF99'})
                            ], style={'backgroundColor': '#1a1a2e', 'border': '1px solid #00FF99'}),
                            dbc.ListGroupItem([
                                html.H6("GET /api/v1/scan-results", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                                html.P("Tarama sonuçlarını getirir", className="text-muted small mb-0", style={'fontSize': '0.8rem'}),
                                html.Code("Authorization: Bearer YOUR_API_KEY", className="small", style={'fontSize': '0.75rem', 'color': '#00FF99'})
                            ], style={'backgroundColor': '#1a1a2e', 'border': '1px solid #00FF99'}),
                            dbc.ListGroupItem([
                                html.H6("POST /api/v1/start-scan", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                                html.P("Yeni tarama başlatır", className="text-muted small mb-0", style={'fontSize': '0.8rem'}),
                                html.Code("Authorization: Bearer YOUR_API_KEY", className="small", style={'fontSize': '0.75rem', 'color': '#00FF99'})
                            ], style={'backgroundColor': '#1a1a2e', 'border': '1px solid #00FF99'}),
                        ], flush=True),
                    ], style={'padding': '1.5rem', 'backgroundColor': '#1a1a2e'})
                ], style={
                    'border': '1px solid #00FF99',
                    'borderRadius': '12px',
                    'boxShadow': '0 0 20px rgba(0,255,153,0.15)'
                })
            ], width=8, className="mx-auto")
        ])
    ], fluid=True, className="mt-4", style={'backgroundColor': '#0F1114'})

# ====================== HESAP AYARLARI SAYFASI ======================
def create_settings_layout():
    if not LIVE_DATA['current_user']:
        return login_layout
        
    return dbc.Container([
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader(html.H4("⚙️ Gelişmiş Hesap Ayarları", className="mb-0", style={
                        'fontSize': '1.2rem',
                        'color': '#00FF99',
                        'textShadow': '0 0 5px rgba(0,255,153,0.3)'
                    })),
                    dbc.CardBody([
                        dbc.Form([
                            dbc.Row([
                                dbc.Col([
                                    dbc.Label("Kullanıcı Adı", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                                    dbc.Input(
                                        id='settings-username',
                                        value=LIVE_DATA['current_user']['username'],
                                        readonly=True,
                                        style={
                                            'fontSize': '0.85rem',
                                            'backgroundColor': '#0F1114',
                                            'border': '1px solid #00FF99',
                                            'color': 'white',
                                            'borderRadius': '8px'
                                        }
                                    ),
                                ], width=6),
                                dbc.Col([
                                    dbc.Label("E-posta Adresi", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                                    dbc.Input(
                                        id='settings-email',
                                        type='email',
                                        value=LIVE_DATA['current_user']['email'],
                                        style={
                                            'fontSize': '0.85rem',
                                            'backgroundColor': '#0F1114',
                                            'border': '1px solid #00FF99',
                                            'color': 'white',
                                            'borderRadius': '8px'
                                        }
                                    ),
                                ], width=6),
                            ]),
                            
                            dbc.Label("Şirket", style={'fontSize': '0.9rem', 'color': '#00FF99'}),
                            dbc.Input(
                                id='settings-company',
                                value=LIVE_DATA['current_user'].get('company', ''),
                                style={
                                    'fontSize': '0.85rem',
                                    'backgroundColor': '#0F1114',
                                    'border': '1px solid #00FF99',
                                    'color': 'white',
                                    'borderRadius': '8px'
                                }
                            ),
                            
                            dbc.Checklist(
                                options=[
                                    {"label": "Kritik güvenlik bulgularında e-posta ile bilgilendir", "value": 1}
                                ],
                                value=[1] if LIVE_DATA['current_user'].get('email_notifications', 1) else [],
                                id="settings-notifications",
                                style={'fontSize': '0.9rem', 'color': '#00FF99'}
                            ),
                            
                            dbc.Button("Ayarları Kaydet", id="save-settings", color="primary", className="mt-3", size="sm", style={
                                'backgroundColor': '#00FF99',
                                'border': 'none',
                                'color': '#0F1114',
                                'fontWeight': 'bold',
                                'borderRadius': '6px'
                            }),
                            html.Div(id='settings-alert', className="mt-3", style={'fontSize': '0.9rem'}),
                        ])
                    ], style={'padding': '1.5rem', 'backgroundColor': '#1a1a2e'})
                ], style={
                    'border': '1px solid #00FF99',
                    'borderRadius': '12px',
                    'boxShadow': '0 0 20px rgba(0,255,153,0.15)'
                })
            ], width=8, className="mx-auto")
        ])
    ], fluid=True, className="mt-4", style={'backgroundColor': '#0F1114'})

# ====================== CALLBACK'LER ======================

# Sayfa yönlendirme
@app.callback(
    Output('page-content', 'children'),
    [Input('url', 'pathname')],
    [State('user-store', 'data')]
)
def display_page(pathname, user_data):
    if not user_data:
        return login_layout
    
    if pathname == '/api':
        return create_api_layout()
    elif pathname == '/settings':
        return create_settings_layout()
    else:
        return create_dashboard_layout()

# Giriş kontrolü
@app.callback(
    [Output('user-store', 'data'),
     Output('url', 'pathname'),
     Output('login-alert', 'children')],
    [Input('login-button', 'n_clicks')],
    [State('login-username', 'value'),
     State('login-password', 'value')]
)
def login_user(n_clicks, username, password):
    if n_clicks is None or n_clicks == 0:
        return dash.no_update, dash.no_update, ""
    
    if not username or not password:
        return dash.no_update, dash.no_update, dbc.Alert("Kullanıcı adı ve şifre gereklidir!", color="danger", style={'fontSize': '0.9rem'})
    
    try:
        # Şifreyi hash'le ve kontrol et
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute('SELECT username, email, role, company, email_notifications FROM users WHERE username = ? AND password_hash = ? AND is_active = 1', 
                       (username, password_hash))
        user = cursor.fetchone()
        conn.close()
        
        if user:
            user_data = {
                'username': user[0],
                'email': user[1],
                'role': user[2],
                'company': user[3],
                'email_notifications': user[4]
            }
            LIVE_DATA['current_user'] = user_data
            
            # Tarama geçmişini yükle
            conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT scan_type, target, created_at, critical_findings 
                FROM scan_history 
                WHERE user_id = (SELECT id FROM users WHERE username = ?)
                ORDER BY created_at DESC LIMIT 1
            ''', (username,))
            last_scan = cursor.fetchone()
            conn.close()
            
            if last_scan:
                LIVE_DATA['last_update'] = datetime.strptime(last_scan[2], '%Y-%m-%d %H:%M:%S')
            
            return user_data, '/', ""
        else:
            return dash.no_update, dash.no_update, dbc.Alert("Geçersiz kullanıcı adı veya şifre!", color="danger", style={'fontSize': '0.9rem'})
    except Exception as e:
        print(f"Giriş hatası: {e}")
        return dash.no_update, dash.no_update, dbc.Alert(f"Sistem hatası: {e}", color="danger", style={'fontSize': '0.9rem'})

# Çıkış işlemi
@app.callback(
    [Output('user-store', 'data', allow_duplicate=True),
     Output('url', 'pathname', allow_duplicate=True)],
    [Input('logout-button', 'n_clicks')],
    prevent_initial_call=True
)
def logout_user(n_clicks):
    if n_clicks:
        LIVE_DATA['current_user'] = None
        return None, '/'
    return dash.no_update, dash.no_update

# Dashboard güncelleme
@app.callback(
    [Output('network-graph', 'figure'),
     Output('device-pie', 'figure'),
     Output('timeline-graph', 'figure'),
     Output('criticality-bar', 'figure'),
     Output('device-table', 'data'),
     Output('stat-devices', 'children'),
     Output('stat-high-risk', 'children'),
     Output('stat-protocols', 'children'),
     Output('stat-last-scan', 'children')],
    [Input('interval-component', 'n_intervals')]
)
def update_dashboard(n):
    data = LIVE_DATA
    
    # İstatistikleri güncelle
    total_devices = len(data["devices"])
    high_risk = len([d for d in data["devices"] if d.get('criticality') == 'high'])
    
    # Protokol sayısını hesapla
    protocols = 0
    if data["devices"]:
        all_protocols = set()
        for d in data["devices"]:
            protocols_value = d.get('protocols', '')
            if isinstance(protocols_value, str):
                if protocols_value:
                    protocols_list = [p.strip() for p in protocols_value.split(',')]
                    all_protocols.update(protocols_list)
            elif isinstance(protocols_value, list):
                all_protocols.update(protocols_value)
        protocols = len(all_protocols)
    
    last_scan = data["last_update"].strftime("%H:%M") if data["last_update"] else "-"
    
    # Varsayılan grafikler
    network_fig = go.Figure()
    device_pie = go.Figure()
    timeline_fig = go.Figure()
    criticality_fig = go.Figure()
    
    # Grafikleri oluştur
    if data["devices"]:
        # Network grafiği
        device_ips = [d.get('ip', 'Unknown') for d in data["devices"]]
        risk_scores = [random.randint(1, 100) for _ in data["devices"]]
        
        network_fig.add_trace(go.Scatter(
            x=list(range(len(data["devices"]))),
            y=risk_scores,
            mode='markers+text',
            marker=dict(
                size=15, 
                color=risk_scores,
                colorscale='RdYlGn_r',
                showscale=True,
                colorbar=dict(title="Risk Skoru")
            ),
            text=[f"Cihaz {i+1}" for i in range(len(data["devices"]))],
            textposition="middle center",
            hovertemplate='<b>%{text}</b><br>IP: %{customdata}<br>Risk Skoru: %{y}<extra></extra>',
            customdata=device_ips
        ))
        network_fig.update_layout(
            title=dict(text="Ağ Cihaz Risk Dağılımı", font=dict(size=14, color='#00FF99')),
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='white', size=10),
            margin=dict(l=20, r=20, t=40, b=20),
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, title=None),
            yaxis=dict(showgrid=True, zeroline=True, showticklabels=True, title="Risk Skoru", gridcolor='#333'),
            showlegend=False
        )
        
        # Pasta grafiği
        vendors = [d.get('vendor', 'Bilinmiyor') for d in data["devices"]]
        vendor_counts = pd.Series(vendors).value_counts()
        device_pie = px.pie(
            values=vendor_counts.values,
            names=vendor_counts.index,
            title="Üretici Dağılımı",
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        device_pie.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(color='white', size=10),
            margin=dict(l=20, r=20, t=40, b=20),
            showlegend=True,
            legend=dict(font=dict(size=9)),
            title_font_color='#00FF99'
        )
        device_pie.update_traces(
            textposition='inside',
            textinfo='percent+label',
            textfont_size=10,
            hoverinfo='label+percent+value'
        )
        
        # Zaman çizelgesi
        if len(data["devices"]) > 0:
            timeline_data = []
            base_time = datetime.now()
            for i, device in enumerate(data["devices"]):
                timeline_data.append({
                    'time': (base_time - timedelta(minutes=random.randint(0, 60))).strftime("%H:%M"),
                    'device': f"Cihaz {i+1}",
                    'risk': random.randint(1, 100),
                    'ip': device.get('ip', 'N/A')
                })
            
            timeline_df = pd.DataFrame(timeline_data)
            timeline_fig = px.scatter(
                timeline_df,
                x='time',
                y='risk',
                title="Risk Zaman Çizelgesi",
                labels={'time': 'Zaman', 'risk': 'Risk Skoru'},
                color='risk',
                color_continuous_scale='RdYlGn_r',
                hover_data=['ip']
            )
            timeline_fig.update_traces(
                marker=dict(size=8, line=dict(width=1, color='DarkSlateGrey')),
                hovertemplate='<b>%{customdata[0]}</b><br>Zaman: %{x}<br>Risk: %{y}<extra></extra>'
            )
            timeline_fig.update_layout(
                paper_bgcolor='rgba(0,0,0,0)',
                plot_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white', size=10),
                margin=dict(l=20, r=20, t=40, b=20),
                xaxis=dict(tickfont=dict(size=8), title_font=dict(size=9), gridcolor='#333'),
                yaxis=dict(tickfont=dict(size=8), title_font=dict(size=9), gridcolor='#333'),
                title_font_color='#00FF99'
            )
        
        # Kritiklik grafiği
        criticalities = [d.get('criticality', 'low') for d in data["devices"]]
        crit_counts = pd.Series(criticalities).value_counts().reindex(['low', 'medium', 'high'], fill_value=0)
        criticality_fig = px.bar(
            x=crit_counts.index,
            y=crit_counts.values,
            title="Risk Seviyeleri",
            color=crit_counts.index,
            color_discrete_map={'high': '#ff4444', 'medium': '#ffaa00', 'low': '#00FF99'},
            labels={'x': 'Risk Seviyesi', 'y': 'Cihaz Sayısı'}
        )
        criticality_fig.update_layout(
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            font=dict(color='white', size=10),
            margin=dict(l=20, r=20, t=40, b=20),
            showlegend=False,
            xaxis=dict(title_font=dict(size=9), tickfont=dict(size=9)),
            yaxis=dict(title_font=dict(size=9), tickfont=dict(size=9), gridcolor='#333'),
            title_font_color='#00FF99'
        )
        criticality_fig.update_traces(
            hovertemplate='<b>%{x}</b><br>Cihaz Sayısı: %{y}<extra></extra>'
        )
    
    # Boş grafikler için varsayılan layout
    empty_layout = dict(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='white', size=12),
        margin=dict(l=20, r=20, t=60, b=20),
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False, title=None),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False, title=None),
        annotations=[dict(
            text="Veri bekleniyor...<br>Tarama başlatın",
            xref="paper", yref="paper",
            x=0.5, y=0.5, xanchor='center', yanchor='middle',
            showarrow=False,
            font=dict(size=14, color='#666')
        )]
    )
    
    if not data["devices"]:
        network_fig.update_layout(**empty_layout, title=dict(text="Ağ Cihaz Risk Dağılımı", font=dict(size=14, color='#00FF99')))
        device_pie.update_layout(**empty_layout, title=dict(text="Üretici Dağılımı", font=dict(size=14, color='#00FF99')))
        timeline_fig.update_layout(**empty_layout, title=dict(text="Risk Zaman Çizelgesi", font=dict(size=14, color='#00FF99')))
        criticality_fig.update_layout(**empty_layout, title=dict(text="Risk Seviyeleri", font=dict(size=14, color='#00FF99')))
    
    return (network_fig, device_pie, timeline_fig, criticality_fig, 
            data["devices"], total_devices, high_risk, protocols, last_scan)

# Tarama başlatma - ORCHESTRATOR ENTEGRASYONLU
@app.callback(
    [Output('scan-status', 'children'),
     Output('active-scan-id', 'data')],
    [Input('start-scan', 'n_clicks')],
    [State('scan-type', 'value'),
     State('target-input', 'value'),
     State('email-input', 'value'),
     State('user-store', 'data')]
)
def start_orchestrated_scan(n_clicks, scan_type, target, email, user_data):
    if n_clicks is None or n_clicks == 0:
        return "", dash.no_update
    
    if not scan_type or not target:
        return dbc.Alert("Lütfen tarama türü ve hedef belirleyin!", color="warning", style={'fontSize': '0.85rem'}), dash.no_update
    
    # E-posta adresini kontrol et veya kullanıcının e-posta adresini kullan
    notification_email = email or user_data.get('email')
    if notification_email:
        user_data['email'] = notification_email
    
    # Orchestrator ile taramayı başlat
    scan_id = orchestrator.analyze_target(scan_type, target, user_data)
    if not scan_id:
        return dbc.Alert("Görev kuyruğa eklenemedi. Orchestrator API'yi kontrol edin.", color="danger", style={'fontSize': '0.85rem'}), dash.no_update

    module_info = orchestrator.modules.get(scan_type, {})
    module_name = module_info.get('description', 'Bilinmeyen Modül')
    
    alert = dbc.Alert(
        [
            html.H6("🚀 Gelişmiş Tarama Başlatıldı", className="alert-heading"),
            html.P(f"Modül: {module_name}"),
            html.P(f"Hedef: {target}"),
            html.P(f"Tarama ID: {scan_id}"),
            html.Hr(),
            html.P("Tarama arka planda çalışıyor... Sonuçlar otomatik güncellenecek.", className="mb-0")
        ],
        color="info",
        style={'fontSize': '0.85rem'}
    )
    
    return alert, scan_id

# Tarama durumu güncelleme
@app.callback(
    Output('scan-progress', 'children'),
    [Input('scan-status-interval', 'n_intervals')],
    [State('active-scan-id', 'data')]
)
def update_scan_progress(n, scan_id):
    if not scan_id:
        return ""
    
    scan_status = orchestrator.get_scan_status(scan_id)
    status = scan_status.get('status', 'unknown')

    if status in {'pending', 'queued'}:
        return dbc.Alert("⏳ Görev sırada bekliyor...", color="info", style={'fontSize': '0.85rem'})
    if status in {'running', 'started'}:
        return dbc.Progress(
            value=100,
            striped=True,
            animated=True,
            style={'height': '8px'},
            label="Tarama devam ediyor..."
        )
    if status == 'success':
        return dbc.Alert("✅ Tarama tamamlandı!", color="success", style={'fontSize': '0.85rem'})
    if status == 'failed' or status == 'error':
        return dbc.Alert(f"❌ Tarama hatası: {scan_status.get('error', 'Bilinmeyen hata')}", color="danger", style={'fontSize': '0.85rem'})
    return ""

# OpenAI asistan entegrasyonu
@app.callback(
    Output('assistant-response', 'children'),
    Input('ask-assistant', 'n_clicks'),
    [State('assistant-prompt', 'value'),
     State('target-input', 'value'),
     State('user-store', 'data')]
)
def run_ai_assistant(n_clicks, prompt, target, user_data):
    if not n_clicks:
        return dash.no_update

    if not prompt:
        return dbc.Alert("Lütfen asistan için bir istek veya soru girin.", color="warning", style={'fontSize': '0.85rem'})

    context = {
        'target': target,
        'summary': {
            'total_devices': LIVE_DATA.get('total_devices', 0),
            'high_risk_devices': LIVE_DATA.get('high_risk_devices', 0)
        },
        'devices': LIVE_DATA.get('devices', [])[:5],
        'notes': f"Kullanıcı: {user_data.get('username')}" if user_data else None
    }

    assistant_result = orchestrator.invoke_assistant(prompt, context)
    if assistant_result.get('error'):
        return dbc.Alert(f"Asistan hatası: {assistant_result['error']}", color="danger", style={'fontSize': '0.85rem'})

    response_text = assistant_result.get('assistant_response', 'Asistan yanıt üretemedi.')
    return dbc.Alert(dcc.Markdown(response_text), color="info", style={'fontSize': '0.9rem'})

# Rapor indirme
@app.callback(
    Output("download-excel", "data"),
    [Input("btn-excel", "n_clicks")],
    [State('user-store', 'data')],
    prevent_initial_call=True
)
def download_excel(n_clicks, user_data):
    if n_clicks:
        scan_info = {
            'scan_type': 'OT/ICS Tarama',
            'target': '192.168.1.0/24',
            'username': user_data['username']
        }
        buffer = report_service.generate_excel_report(LIVE_DATA, scan_info)
        return dcc.send_bytes(buffer.getvalue(), f"varux_rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

@app.callback(
    Output("download-pdf", "data"),
    [Input("btn-pdf", "n_clicks")],
    [State('user-store', 'data')],
    prevent_initial_call=True
)
def download_pdf(n_clicks, user_data):
    if n_clicks:
        scan_info = {
            'scan_type': 'OT/ICS Tarama',
            'target': '192.168.1.0/24', 
            'username': user_data['username']
        }
        pdf_data = report_service.generate_pdf_report(LIVE_DATA, scan_info)
        return dcc.send_bytes(pdf_data, f"varux_rapor_{datetime.now().strftime('%Y%m%d_%H%M')}.html")

# API anahtar yönetimi
@app.callback(
    Output('api-key-input', 'value'),
    [Input('generate-api-key', 'n_clicks')],
    [State('user-store', 'data')],
    prevent_initial_call=True
)
def generate_api_key(n_clicks, user_data):
    if n_clicks:
        new_api_key = secrets.token_hex(32)
        conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute('UPDATE users SET api_key = ? WHERE username = ?', (new_api_key, user_data['username']))
        conn.commit()
        conn.close()
        return new_api_key
    return dash.no_update

# Hesap ayarlarını kaydetme
@app.callback(
    Output('settings-alert', 'children'),
    [Input('save-settings', 'n_clicks')],
    [State('settings-email', 'value'),
     State('settings-company', 'value'),
     State('settings-notifications', 'value'),
     State('user-store', 'data')],
    prevent_initial_call=True
)
def save_settings(n_clicks, email, company, notifications, user_data):
    if n_clicks:
        try:
            conn = sqlite3.connect('varux_enterprise.db', check_same_thread=False)
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE users SET email = ?, company = ?, email_notifications = ? WHERE username = ?',
                (email, company, 1 if notifications else 0, user_data['username'])
            )
            conn.commit()
            conn.close()
            
            # Kullanıcı verisini güncelle
            LIVE_DATA['current_user']['email'] = email
            LIVE_DATA['current_user']['company'] = company
            
            return dbc.Alert("✅ Ayarlar başarıyla kaydedildi!", color="success", style={'fontSize': '0.9rem'})
        except Exception as e:
            return dbc.Alert(f"❌ Ayarlar kaydedilemedi: {e}", color="danger", style={'fontSize': '0.9rem'})
    return ""

# ====================== API ENDPOINT'LERİ ======================
@server.route('/api/v1/devices', methods=['GET'])
def api_get_devices():
    api_key = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = api_service.validate_api_key(api_key)
    
    if not user:
        return jsonify({'error': 'Geçersiz API anahtarı'}), 401
    
    # API isteğini logla
    api_service.log_api_request(api_key, '/api/v1/devices', request.remote_addr, request.headers.get('User-Agent'))
    
    return jsonify({
        'status': 'success',
        'data': LIVE_DATA['devices'],
        'total_devices': len(LIVE_DATA['devices']),
        'last_update': LIVE_DATA['last_update'].isoformat() if LIVE_DATA['last_update'] else None
    })

@server.route('/api/v1/scan-results', methods=['GET'])
def api_get_scan_results():
    api_key = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = api_service.validate_api_key(api_key)
    
    if not user:
        return jsonify({'error': 'Geçersiz API anahtarı'}), 401
    
    api_service.log_api_request(api_key, '/api/v1/scan-results', request.remote_addr, request.headers.get('User-Agent'))
    
    return jsonify({
        'status': 'success',
        'data': LIVE_DATA
    })

@server.route('/api/v1/assistant', methods=['POST'])
def api_ai_assistant():
    api_key = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = api_service.validate_api_key(api_key)

    if not user:
        return jsonify({'error': 'Geçersiz API anahtarı'}), 401

    payload = request.get_json() or {}
    prompt = payload.get('prompt')
    context = payload.get('context', {})

    if not prompt:
        return jsonify({'error': 'prompt gereklidir'}), 400

    context.setdefault('summary', {
        'total_devices': LIVE_DATA.get('total_devices', 0),
        'high_risk_devices': LIVE_DATA.get('high_risk_devices', 0)
    })
    context.setdefault('devices', LIVE_DATA.get('devices', [])[:5])
    context.setdefault('notes', f"API kullanıcı: {user[0]}")

    api_service.log_api_request(api_key, '/api/v1/assistant', request.remote_addr, request.headers.get('User-Agent'))
    assistant_result = orchestrator.invoke_assistant(prompt, context)

    status_code = 200 if not assistant_result.get('error') else 400
    return jsonify(assistant_result), status_code

@server.route('/api/v1/start-scan', methods=['POST'])
def api_start_scan():
    api_key = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = api_service.validate_api_key(api_key)
    
    if not user:
        return jsonify({'error': 'Geçersiz API anahtarı'}), 401
    
    data = request.get_json()
    scan_type = data.get('scan_type')
    target = data.get('target')
    
    if not scan_type or not target:
        return jsonify({'error': 'scan_type ve target gereklidir'}), 400
    
    # Kullanıcı verisini oluştur
    user_data = {
        'username': user[0],
        'email': f"{user[0]}@api.user",
        'email_notifications': False
    }
    
    # Orchestrator ile tarama başlat
    scan_id = orchestrator.analyze_target(scan_type, target, user_data)
    
    return jsonify({
        'status': 'success',
        'message': f'{scan_type} taraması başlatıldı',
        'scan_id': scan_id
    })

# ====================== UYGULAMAYI BAŞLAT ======================
if __name__ == '__main__':
    print("🚀 VARUX Enterprise Security Platform v6.0 - Orchestrator başlatılıyor...")
    print("📍 Dashboard: http://127.0.0.1:8050")
    print("🔑 Demo Giriş: admin / admin123")
    print("🎯 Orchestrator: Aktif")
    print("📧 E-posta Servisi: Aktif")
    print("📊 Rapor Servisi: Aktif")
    print("🔌 API Servisi: Aktif")
    print("💾 Veritabanı: Başlatıldı")
    print("⚡ Modül Entegrasyonu: Tamamlandı")

    app.run(debug=True, port=8050, host='0.0.0.0')
